from django.contrib.auth.tokens import PasswordResetTokenGenerator
import six
import openpyxl
from openpyxl.styles import Alignment, Font
from openpyxl.drawing.image import Image
from django.http import HttpResponse
from datetime import datetime
from twilio.rest import Client
from django.conf import settings
from django.db.models.functions import TruncDay
from .models import Booking, User
from django.db.models import Sum, Count
import os
import subprocess
from datetime import datetime
from django.conf import settings

class TokenGenerator(PasswordResetTokenGenerator):

    def _make_hash_value(self, user, timestamp):
        return (six.text_type(user.pk)+six.text_type(timestamp)+six.text_type(user.is_email_verified))

generate_token = TokenGenerator()

def get_twilio_client():
    return Client(settings.TWILIO_ACCOUNT_SID, settings.TWILIO_AUTH_TOKEN)

def send_sms(to, message):
    client = get_twilio_client()
    message = client.messages.create(
        body=message,
        from_=settings.TWILIO_PHONE_NUMBER,
        to=to
    )
    return message.sid

def get_user_activity_report():
    # Example: Get user registration counts by date
    data = User.objects.annotate(date_registered=TruncDay('date_joined')).values('date_registered').annotate(count=Count('id')).order_by('date_registered')
    return data

def get_sales_report():
    # Example: Sum of payments received by date
    data = Booking.objects.annotate(date=TruncDay('booking_date')).values('date').annotate(total_sales=Sum('payment_amount')).order_by('date')
    return data

def generate_excel(queryset):
    # Create a workbook and select the active worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Booking Report"

    # Set the column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 30
    ws.column_dimensions['I'].width = 40

    # Add company logo
    logo_path = os.path.join(settings.MEDIA_ROOT, 'movie_images', 'logo2.jpg')  # Replace 'logo.jpg' with your logo filename
    if os.path.exists(logo_path):
        logo = Image(logo_path)
        logo.width = 100  # Set the width of the logo
        logo.height = 100  # Set the height of the logo
        ws.add_image(logo, 'A1')
    else:
        ws['A1'] = "Logo not found"
    
    # Merge cells for company details
    ws.merge_cells('B1:E1')
    ws.merge_cells('B2:E2')
    ws.merge_cells('B3:E3')
    ws.merge_cells('B4:E4')
    ws.merge_cells('B5:E5')

    # Add company details
    ws['B1'] = "Company Name: WatchIt"
    ws['B2'] = "Address: 123 Laucala Bay, Suva"
    ws['B3'] = "Phone: +(679) 999-7890"
    ws['B4'] = "Email: info@watchit.com"

    # Add report title
    ws['A6'] = "Booking Report"
    ws['A7'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A8'] = ""

    # Add the table headers
    headers = ['ID', 'User', 'Movie', 'Cinema Hall', 'Showtime', 'Booking Date', 'Payment Amount', 'Seats']
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=9, column=col_num)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Add the data rows
    row_num = 10
    for booking in queryset:
        booking.seat_labels = ', '.join(seat.seat_label for seat in booking.seats.all())
        showtime_display = booking.showtime.showtime.strftime("%Y-%m-%d %H:%M:%S")  # Format the showtime to a string
        row = [
            booking.id,
            booking.user.username if booking.user else 'Guest User',
            booking.movie.title,
            booking.cinema_hall.cinema_type,
            showtime_display,  # Convert showtime to string
            booking.booking_date.strftime("%Y-%m-%d %H:%M:%S"),  # Convert booking_date to string
            f"${booking.payment_amount:.2f}",
            booking.seat_labels
        ]
        for col_num, value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal="center", vertical="center")
        row_num += 1

    # Save the workbook to a response object
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sales_report.xlsx"'
    wb.save(response)

    return response

# def backup_database():
#     USER = 'karan'
#     HOST = 'localhost'
#     PORT = '5432'
#     DB_NAME = 'WatchIt'
#     BACKUP_DIR = os.path.join(settings.BASE_DIR, 'WatchIt-DB-Backup')
#     if not os.path.exists(BACKUP_DIR):
#         os.makedirs(BACKUP_DIR)
#     TIMESTAMP = datetime.now().strftime("%Y%m%d%H%M%S")
#     BACKUP_FILE = os.path.join(BACKUP_DIR, f'{DB_NAME}_{TIMESTAMP}.sql')

#     try:
#         os.system(f'pg_dump -U {USER} -h {HOST} -p {PORT} {DB_NAME} > {BACKUP_FILE}')
#         return BACKUP_FILE
#     except Exception as e:
#         print(f"Error during backup: {e}")
#         return None

# def restore_database(file_path):
#     USER = 'karan'
#     HOST = 'localhost'
#     PORT = '5432'
#     DB_NAME = 'WatchIt'

#     try:
#         os.system(f'psql -U {USER} -h {HOST} -p {PORT} {DB_NAME} < {file_path}')
#     except Exception as e:
#         print(f"Error during restore: {e}")


# def backup_database():
#     USER = 'karan'
#     HOST = 'localhost'
#     PORT = '5432'
#     DB_NAME = 'WatchIt'
#     BACKUP_DIR = os.path.join(settings.BASE_DIR, 'WatchIt-DB-Backup')
#     if not os.path.exists(BACKUP_DIR):
#         os.makedirs(BACKUP_DIR)
#     TIMESTAMP = datetime.now().strftime("%Y%m%d%H%M%S")
#     BACKUP_FILE = os.path.join(BACKUP_DIR, f'{DB_NAME}_{TIMESTAMP}.sql')

#     try:
#         command = f'pg_dump -U {USER} -h {HOST} -p {PORT} {DB_NAME} > {BACKUP_FILE}'
#         result = os.system(command)
#         if result != 0:
#             raise Exception(f"Backup command failed with exit code {result}")
#         print("Database backup successful")
#         return BACKUP_FILE
#     except Exception as e:
#         print(f"Error during backup: {e}")
#         return None

# def restore_database(file_path):
#     USER = 'karan'
#     HOST = 'localhost'
#     PORT = '5432'
#     DB_NAME = 'WatchIt'

#     os.system(f'psql -U {USER} -h {HOST} -p {PORT} {DB_NAME} < {file_path}')


import os
import subprocess
from datetime import datetime
from django.conf import settings

def backup_database():
    USER = settings.DATABASES['default']['USER']
    HOST = settings.DATABASES['default']['HOST']
    PORT = settings.DATABASES['default']['PORT']
    DB_NAME = settings.DATABASES['default']['NAME']
    BACKUP_DIR = os.path.join(settings.BASE_DIR, 'WatchIt-DB-Backup')

    if not os.path.exists(BACKUP_DIR):
        os.makedirs(BACKUP_DIR)

    TIMESTAMP = datetime.now().strftime("%Y%m%d%H%M%S")
    BACKUP_FILE = os.path.join(BACKUP_DIR, f'{DB_NAME}_data_{TIMESTAMP}.sql')

    command = ['pg_dump', '-U', USER, '-h', HOST, '-p', PORT, '-F', 'p', '-f', BACKUP_FILE, DB_NAME]

    try:
        result = subprocess.run(command, check=True, stderr=subprocess.PIPE, text=True)
        print("Database backup successful")
        return BACKUP_FILE
    except subprocess.CalledProcessError as e:
        print(f"Error during backup: {e.stderr}")
        return None
    except Exception as e:
        print(f"Unexpected error: {e}")
        return None


# def restore_database(file_path):
#     USER = settings.DATABASES['default']['USER']
#     HOST = settings.DATABASES['default']['HOST']
#     PORT = settings.DATABASES['default']['PORT']
#     DB_NAME = settings.DATABASES['default']['NAME']

#     drop_tables_command = f"""
#     psql -U {USER} -h {HOST} -p {PORT} -d {DB_NAME} -c "
#     DO $$ DECLARE
#     r RECORD;
#     BEGIN
#     FOR r IN (SELECT tablename FROM pg_tables WHERE schemaname = current_schema()) LOOP
#         EXECUTE 'DROP TABLE IF EXISTS ' || quote_ident(r.tablename) || ' CASCADE';
#     END LOOP;
#     END $$;
#     "
#     """
#     try:
#         print("Dropping existing tables...")
#         subprocess.run(drop_tables_command, shell=True, check=True, stderr=subprocess.PIPE, text=True)
#         print("Existing tables dropped successfully")
#     except subprocess.CalledProcessError as e:
#         print(f"Error dropping tables: {e.stderr}")
#         return

#     command = ['psql', '-U', USER, '-h', HOST, '-p', PORT, '-d', DB_NAME, '-f', file_path]
#     try:
#         print("Restoring the database...")
#         result = subprocess.run(command, check=True, text=True, stderr=subprocess.PIPE, stdout=subprocess.PIPE)
#         print("Database restore successful")
#         print(result.stdout)
#     except subprocess.CalledProcessError as e:
#         print(f"Error during restore: {e.stderr}")
#     except Exception as e:
#         print(f"Unexpected error: {e}")

# # Example usage:
# # backup_file = backup_database()
# # if backup_file:
# #     restore_database(backup_file)

import subprocess

def drop_all_tables_and_sequences(db_name='WatchIt', user='karan', host='localhost', port='5432'):
    tables = [
        "WatchIt_cinemahall", "WatchIt_movie", "WatchIt_showtime", "WatchIt_user", "WatchIt_booking",
        "WatchIt_seat", "WatchIt_booking_seats", "WatchIt_feedback", "WatchIt_tag", "WatchIt_movie_tags",
        "auth_group", "WatchIt_user_groups", "django_content_type", "auth_permission", "WatchIt_user_user_permissions",
        "auth_group_permissions", "django_admin_log", "django_migrations", "django_session"
    ]
    for table in tables:
        command = f'psql -U {user} -h {host} -p {port} -d {db_name} -c "DROP TABLE IF EXISTS \\"{table}\\" CASCADE;"'
        try:
            result = subprocess.run(command, shell=True, check=True, stderr=subprocess.PIPE, stdout=subprocess.PIPE, text=True)
            print(f"Dropped table: {table}")
            print("STDOUT:", result.stdout)
            print("STDERR:", result.stderr)
        except subprocess.CalledProcessError as e:
            print(f"Error dropping table {table}: {e.stderr}")

    sequences = [
        "WatchIt_cinemahall_id_seq", "WatchIt_movie_id_seq", "WatchIt_showtime_id_seq", "WatchIt_user_id_seq", "WatchIt_booking_id_seq",
        "WatchIt_seat_id_seq", "WatchIt_booking_seats_id_seq", "WatchIt_feedback_id_seq", "WatchIt_tag_id_seq", "WatchIt_movie_tags_id_seq",
        "auth_group_id_seq", "WatchIt_user_groups_id_seq", "django_content_type_id_seq", "auth_permission_id_seq", "WatchIt_user_user_permissions_id_seq",
        "auth_group_permissions_id_seq", "django_admin_log_id_seq", "django_migrations_id_seq", "django_session_id_seq"
    ]
    for sequence in sequences:
        command = f'psql -U {user} -h {host} -p {port} -d {db_name} -c "DROP SEQUENCE IF EXISTS \\"{sequence}\\" CASCADE;"'
        try:
            result = subprocess.run(command, shell=True, check=True, stderr=subprocess.PIPE, stdout=subprocess.PIPE, text=True)
            print(f"Dropped sequence: {sequence}")
            print("STDOUT:", result.stdout)
            print("STDERR:", result.stderr)
        except subprocess.CalledProcessError as e:
            print(f"Error dropping sequence {sequence}: {e.stderr}")

def restore_database(file_path, db_name='WatchIt', user='karan', host='localhost', port='5432'):
    # Step 1: Drop all tables and sequences
    drop_all_tables_and_sequences(db_name, user, host, port)

    # Step 2: Disable constraints
    disable_constraints_command = f'psql -U {user} -h {host} -p {port} -d {db_name} -c "SET session_replication_role = replica;"'
    enable_constraints_command = f'psql -U {user} -h {host} -p {port} -d {db_name} -c "SET session_replication_role = DEFAULT;"'

    # Step 3: Restore the database
    restore_command = f'psql -U {user} -h {host} -p {port} -d {db_name} -f {file_path}'

    try:
        print("Disabling constraints...")
        disable_result = subprocess.run(disable_constraints_command, shell=True, check=True, stderr=subprocess.PIPE, text=True)
        print("Constraints disabled")
        print(disable_result.stderr)

        print("Restoring the database...")
        restore_result = subprocess.run(restore_command, shell=True, check=True, text=True, stderr=subprocess.PIPE, stdout=subprocess.PIPE)
        print("Database restore output:")
        print(restore_result.stdout)
        print("Database restore errors:")
        print(restore_result.stderr)

        print("Re-enabling constraints...")
        enable_result = subprocess.run(enable_constraints_command, shell=True, check=True, stderr=subprocess.PIPE, text=True)
        print("Constraints re-enabled")
        print(enable_result.stderr)

    except subprocess.CalledProcessError as e:
        print(f"Error during restore: {e.stderr}")
    except Exception as e:
        print(f"Unexpected error: {e}")

